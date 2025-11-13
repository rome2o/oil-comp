#!/usr/bin/env python3
"""
SEO Analysis Data Preparation Script
====================================

This script handles Phase 1 of the SEO analysis project:
- Environment setup
- Data inventory and validation
- Initial data quality assessment

Author: GitHub Copilot Assistant
Date: October 22, 2025
"""

import os
import pandas as pd
import numpy as np
from pathlib import Path
import openpyxl
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

def create_project_structure():
    """Create the analysis workspace structure"""
    directories = [
        'analysis/notebooks',
        'analysis/processed_data', 
        'analysis/visualizations',
        'analysis/reports',
        'outputs'
    ]
    
    print("Creating project structure...")
    for directory in directories:
        Path(directory).mkdir(parents=True, exist_ok=True)
        print(f"✓ Created: {directory}/")
    
    print("✓ Project structure created successfully")

def inventory_data_files():
    """Inventory and validate all available data files"""
    
    data_inventory = {
        'google_analytics': [],
        'google_search_console': [],
        'ahrefs': []
    }
    
    # Check Google Analytics files
    ga_path = Path('google_analytics')
    if ga_path.exists():
        for file in ga_path.glob('*.xlsx'):
            data_inventory['google_analytics'].append({
                'filename': file.name,
                'path': str(file),
                'size_mb': round(file.stat().st_size / (1024*1024), 2),
                'modified': datetime.fromtimestamp(file.stat().st_mtime)
            })
    
    # Check Google Search Console files
    gsc_path = Path('google_search_console')
    if gsc_path.exists():
        for file in gsc_path.glob('*.xlsx'):
            data_inventory['google_search_console'].append({
                'filename': file.name,
                'path': str(file),
                'size_mb': round(file.stat().st_size / (1024*1024), 2),
                'modified': datetime.fromtimestamp(file.stat().st_mtime)
            })
    
    # Check Ahrefs files
    ahrefs_path = Path('ahrefs')
    if ahrefs_path.exists():
        for file in ahrefs_path.glob('*'):
            if file.suffix in ['.csv', '.pdf']:
                data_inventory['ahrefs'].append({
                    'filename': file.name,
                    'path': str(file),
                    'type': file.suffix[1:],  # Remove the dot
                    'size_mb': round(file.stat().st_size / (1024*1024), 2),
                    'modified': datetime.fromtimestamp(file.stat().st_mtime)
                })
    
    return data_inventory

def validate_excel_files(inventory):
    """Validate Excel files can be opened and check structure"""
    
    validation_results = {}
    
    # Validate Google Analytics files
    for file_info in inventory['google_analytics']:
        try:
            wb = openpyxl.load_workbook(file_info['path'])
            sheet_names = wb.sheetnames
            
            # Try to read with pandas to check structure
            df = pd.read_excel(file_info['path'], sheet_name=0)
            
            validation_results[file_info['filename']] = {
                'status': 'valid',
                'sheets': sheet_names,
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns)[:10]  # First 10 columns
            }
            
        except Exception as e:
            validation_results[file_info['filename']] = {
                'status': 'error',
                'error': str(e)
            }
    
    # Validate Google Search Console files  
    for file_info in inventory['google_search_console']:
        try:
            wb = openpyxl.load_workbook(file_info['path'])
            sheet_names = wb.sheetnames
            
            df = pd.read_excel(file_info['path'], sheet_name=0)
            
            validation_results[file_info['filename']] = {
                'status': 'valid', 
                'sheets': sheet_names,
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns)[:10]
            }
            
        except Exception as e:
            validation_results[file_info['filename']] = {
                'status': 'error',
                'error': str(e)
            }
    
    return validation_results

def validate_csv_files(inventory):
    """Validate CSV files from Ahrefs"""
    
    csv_validation = {}
    
    for file_info in inventory['ahrefs']:
        if file_info['type'] == 'csv':
            try:
                df = pd.read_csv(file_info['path'])
                
                csv_validation[file_info['filename']] = {
                    'status': 'valid',
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': list(df.columns)[:10],
                    'date_columns': [col for col in df.columns if 'date' in col.lower()]
                }
                
            except Exception as e:
                csv_validation[file_info['filename']] = {
                    'status': 'error',
                    'error': str(e)
                }
    
    return csv_validation

def generate_data_quality_report(inventory, excel_validation, csv_validation):
    """Generate comprehensive data quality report"""
    
    report = []
    report.append("# SEO Analysis Data Quality Report")
    report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report.append("")
    
    # Summary
    total_files = (len(inventory['google_analytics']) + 
                  len(inventory['google_search_console']) + 
                  len(inventory['ahrefs']))
    
    report.append("## Summary")
    report.append(f"- Total files found: {total_files}")
    report.append(f"- Google Analytics files: {len(inventory['google_analytics'])}")
    report.append(f"- Google Search Console files: {len(inventory['google_search_console'])}")
    report.append(f"- Ahrefs files: {len(inventory['ahrefs'])}")
    report.append("")
    
    # Google Analytics Details
    report.append("## Google Analytics Files")
    for file_info in inventory['google_analytics']:
        filename = file_info['filename']
        report.append(f"### {filename}")
        report.append(f"- Size: {file_info['size_mb']} MB")
        report.append(f"- Modified: {file_info['modified']}")
        
        if filename in excel_validation:
            validation = excel_validation[filename]
            if validation['status'] == 'valid':
                report.append(f"- Status: ✓ Valid")
                report.append(f"- Rows: {validation['rows']:,}")
                report.append(f"- Columns: {validation['columns']}")
                report.append(f"- Sheets: {', '.join(validation['sheets'])}")
                report.append(f"- Sample columns: {', '.join(validation['column_names'])}")
            else:
                report.append(f"- Status: ❌ Error - {validation['error']}")
        report.append("")
    
    # Google Search Console Details
    report.append("## Google Search Console Files")
    for file_info in inventory['google_search_console']:
        filename = file_info['filename']
        report.append(f"### {filename}")
        report.append(f"- Size: {file_info['size_mb']} MB")
        report.append(f"- Modified: {file_info['modified']}")
        
        if filename in excel_validation:
            validation = excel_validation[filename]
            if validation['status'] == 'valid':
                report.append(f"- Status: ✓ Valid")
                report.append(f"- Rows: {validation['rows']:,}")
                report.append(f"- Columns: {validation['columns']}")
                report.append(f"- Sheets: {', '.join(validation['sheets'])}")
                report.append(f"- Sample columns: {', '.join(validation['column_names'])}")
            else:
                report.append(f"- Status: ❌ Error - {validation['error']}")
        report.append("")
    
    # Ahrefs Details
    report.append("## Ahrefs Files")
    for file_info in inventory['ahrefs']:
        filename = file_info['filename']
        report.append(f"### {filename}")
        report.append(f"- Type: {file_info['type'].upper()}")
        report.append(f"- Size: {file_info['size_mb']} MB") 
        report.append(f"- Modified: {file_info['modified']}")
        
        if file_info['type'] == 'csv' and filename in csv_validation:
            validation = csv_validation[filename]
            if validation['status'] == 'valid':
                report.append(f"- Status: ✓ Valid CSV")
                report.append(f"- Rows: {validation['rows']:,}")
                report.append(f"- Columns: {validation['columns']}")
                report.append(f"- Sample columns: {', '.join(validation['column_names'])}")
                if validation['date_columns']:
                    report.append(f"- Date columns: {', '.join(validation['date_columns'])}")
            else:
                report.append(f"- Status: ❌ Error - {validation['error']}")
        elif file_info['type'] == 'pdf':
            report.append(f"- Status: 📄 PDF (manual extraction required)")
        report.append("")
    
    return '\n'.join(report)

def main():
    """Main execution function for Phase 1"""
    
    print("=== SEO Analysis Phase 1: Environment Setup ===")
    print()
    
    # Task 1.1: Create project structure
    print("Task 1.1: Creating project structure...")
    create_project_structure()
    print()
    
    # Task 1.2: Data inventory
    print("Task 1.2: Performing data inventory...")
    inventory = inventory_data_files()
    
    print("Data files found:")
    for source, files in inventory.items():
        print(f"  {source}: {len(files)} files")
    print()
    
    # Task 1.3: Data validation
    print("Task 1.3: Validating data files...")
    
    print("Validating Excel files...")
    excel_validation = validate_excel_files(inventory)
    
    print("Validating CSV files...")
    csv_validation = validate_csv_files(inventory)
    
    # Generate report
    print("Generating data quality report...")
    report = generate_data_quality_report(inventory, excel_validation, csv_validation)
    
    # Save report
    with open('analysis/reports/data_quality_report.md', 'w') as f:
        f.write(report)
    
    print("✓ Data quality report saved to: analysis/reports/data_quality_report.md")
    print()
    
    print("=== Phase 1 Complete ===")
    print("Next steps:")
    print("1. Review the data quality report")
    print("2. Begin Phase 2: Data extraction and standardization")
    print("3. Run: python data_preparation.py --phase 2")

if __name__ == "__main__":
    main()